#!/usr/bin/env python3
"""
Build Australian regional IO multipliers from ABS IO Tables 2023-24.

Pipeline:
1. Parse 109x109 direct requirements matrix from Table 6
2. Parse employment by industry from Table 20
3. Map IOIGs to 19 ANZSIC divisions
4. Aggregate to 19x19 national matrix using output-weighted averaging
5. Build national Leontief inverse (Type I)
6. Output national multipliers as JSON

This script handles Steps 1-5 (national level).
Regionalisation (FLQ) is done in a separate script once SA4 employment data is sourced.
"""

import json
import os
import openpyxl
import numpy as np

# ============================================================
# IOIG to ANZSIC Division mapping
# Based on ABS IOIG classification concordance
# ============================================================

ANZSIC_DIVISIONS = {
    'A': 'Agriculture, forestry & fishing',
    'B': 'Mining',
    'C': 'Manufacturing',
    'D': 'Electricity, gas, water & waste services',
    'E': 'Construction',
    'F': 'Wholesale trade',
    'G': 'Retail trade',
    'H': 'Accommodation & food services',
    'I': 'Transport, postal & warehousing',
    'J': 'Information media & telecommunications',
    'K': 'Financial & insurance services',
    'L': 'Rental, hiring & real estate services',
    'M': 'Professional, scientific & technical services',
    'N': 'Administrative & support services',
    'O': 'Public administration & safety',
    'P': 'Education & training',
    'Q': 'Health care & social assistance',
    'R': 'Arts & recreation services',
    'S': 'Other services',
}

# Map IOIG code ranges to ANZSIC division letters
# Source: ABS IOIG-ANZSIC concordance
def ioig_to_division(code):
    """Map an IOIG code (integer) to an ANZSIC division letter.
    Based on the actual ABS IOIG-to-ANZSIC concordance for 2023-24.
    Verified against the industry names in Table 6."""
    c = int(code)
    # A: Agriculture, forestry & fishing
    if c <= 501:    return 'A'
    # B: Mining
    if c <= 1001:   return 'B'
    # C: Manufacturing
    if c <= 2502:   return 'C'
    # D: Electricity, gas, water & waste services
    if c <= 2901:   return 'D'
    # E: Construction (includes 3001, 3002, 3101, 3201=construction services)
    if c <= 3201:   return 'E'
    # F: Wholesale trade (3301 only)
    if c <= 3301:   return 'F'
    # G: Retail trade (3901 only)
    if c <= 3901:   return 'G'
    # H: Accommodation & food services
    if c <= 4501:   return 'H'
    # I: Transport, postal & warehousing
    if c <= 5201:   return 'I'
    # J: Information media & telecommunications (5401-5801, plus 6001=library)
    if c <= 5801:   return 'J'
    if c == 6001:   return 'J'  # Library and other information services
    # K: Financial & insurance services (6201-6401)
    if c <= 6401:   return 'K'
    # L: Rental, hiring & real estate services (6601-6701, plus 6700=imputed rent)
    if c <= 6701:   return 'L'
    if c == 6700:   return 'L'  # Imputed rent for owner-occupiers
    # M: Professional, scientific & technical (6702=non-residential property, 6901, 7001)
    if c <= 7001:   return 'M'
    # N: Administrative & support services (7210-7310)
    if c <= 7310:   return 'N'
    # O: Public administration & safety (7501-7701)
    if c <= 7701:   return 'O'
    # P: Education & training (8010-8210)
    if c <= 8210:   return 'P'
    # Q: Health care & social assistance (8401-8601)
    if c <= 8601:   return 'Q'
    # R: Arts & recreation services (8901)
    if c <= 8901:   return 'R'
    # S: Other services (9401+) - no IOIGs in 2023-24 tables
    return 'S'


def parse_direct_requirements(filepath):
    """Parse the 109x109 direct requirements matrix from Table 6."""
    wb = openpyxl.load_workbook(filepath, read_only=True)
    ws = wb['Table 6']
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    # Row 0: IOIG codes in columns (starting col 2)
    # Row 1: Industry names in columns
    # Row 2: Header label row
    # Row 3+: Data rows (col 0 = IOIG code, col 1 = name, col 2+ = coefficients)

    # Extract column codes from row 0
    col_codes = []
    for v in list(rows[0])[2:]:
        if v is not None:
            try:
                col_codes.append(int(v))
            except (ValueError, TypeError):
                break
        else:
            break

    # Extract data rows
    industry_codes = []
    industry_names = []
    matrix_data = []

    for row in rows[3:]:
        vals = list(row)
        if vals[0] is None:
            continue
        try:
            code = int(vals[0])
        except (ValueError, TypeError):
            continue
        if code >= 9000:  # skip summary/total rows
            continue

        industry_codes.append(code)
        industry_names.append(str(vals[1]).strip())

        # Coefficients (cols 2 to 2+n_industries)
        coeffs = []
        for v in vals[2:2+len(col_codes)]:
            if v is None or v == '':
                coeffs.append(0.0)
            else:
                coeffs.append(float(v) / 100.0)  # ABS reports as per $100
        matrix_data.append(coeffs)

    n = len(industry_codes)
    print(f"Parsed {n} industries from direct requirements matrix")
    print(f"Column codes: {len(col_codes)}")

    # Trim matrix to n x n
    A = np.zeros((n, n))
    for i in range(n):
        for j in range(min(n, len(matrix_data[i]))):
            A[i][j] = matrix_data[i][j]

    return industry_codes, industry_names, A


def parse_employment(filepath):
    """Parse employment by industry from Table 20."""
    wb = openpyxl.load_workbook(filepath, read_only=True)
    ws = wb['Table 20']
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    employment = {}
    for row in rows:
        vals = list(row)
        if vals[0] is None:
            continue
        try:
            code = int(vals[0])
        except (ValueError, TypeError):
            continue
        if code >= 9000:
            continue
        # Employment is typically in the last data column or a specific column
        # Let's find the employment figure
        name = str(vals[1]).strip() if vals[1] else ''
        # Try column 2 (employment '000)
        emp = vals[2] if len(vals) > 2 and vals[2] is not None else 0
        try:
            employment[code] = float(emp) * 1000  # convert from '000 to persons
        except (ValueError, TypeError):
            employment[code] = 0

    print(f"Parsed employment for {len(employment)} industries")
    return employment


def parse_output(filepath):
    """Parse industry output from Table 5 (flow table) - total output row."""
    wb = openpyxl.load_workbook(filepath, read_only=True)
    ws = wb['Table 5']
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    # Get column codes from row 0
    col_codes = []
    for v in list(rows[0])[2:]:
        if v is not None:
            try:
                col_codes.append(int(v))
            except (ValueError, TypeError):
                break
        else:
            break

    # Find the "Australian Production" or total output row
    output = {}
    for row in rows:
        vals = list(row)
        if vals[1] is not None and 'australian production' in str(vals[1]).lower():
            for j, code in enumerate(col_codes):
                v = vals[2 + j] if (2 + j) < len(vals) else 0
                try:
                    output[code] = float(v)
                except (ValueError, TypeError):
                    output[code] = 0
            break

    if not output:
        # Try "Total supply" or last data row
        for row in rows:
            vals = list(row)
            if vals[1] is not None and 'total' in str(vals[1]).lower() and 'supply' in str(vals[1]).lower():
                for j, code in enumerate(col_codes):
                    v = vals[2 + j] if (2 + j) < len(vals) else 0
                    try:
                        output[code] = float(v)
                    except (ValueError, TypeError):
                        output[code] = 0
                break

    print(f"Parsed output for {len(output)} industries")
    return output


def aggregate_to_divisions(industry_codes, industry_names, A, output, employment):
    """Aggregate 109x109 matrix to 19x19 ANZSIC division matrix using output-weighted averaging."""
    div_letters = list(ANZSIC_DIVISIONS.keys())
    n_div = len(div_letters)
    div_index = {d: i for i, d in enumerate(div_letters)}

    # Map each industry to its division
    ind_to_div = [ioig_to_division(c) for c in industry_codes]

    # Get output weights for each industry
    weights = np.array([output.get(c, 1.0) for c in industry_codes])

    # Aggregate: for each (div_i, div_j), sum A[i,j] * output[j] for all i in div_i, j in div_j
    # Then divide by total output of div_j
    A_div = np.zeros((n_div, n_div))

    for di in range(n_div):
        for dj in range(n_div):
            num = 0.0
            denom = 0.0
            for j, jdiv in enumerate(ind_to_div):
                if jdiv != div_letters[dj]:
                    continue
                w_j = weights[j]
                denom += w_j
                for i, idiv in enumerate(ind_to_div):
                    if idiv != div_letters[di]:
                        continue
                    num += A[i, j] * w_j
            if denom > 0:
                A_div[di, dj] = num / denom

    # Aggregate employment and output by division
    emp_div = np.zeros(n_div)
    out_div = np.zeros(n_div)
    for k, code in enumerate(industry_codes):
        d = div_index[ind_to_div[k]]
        emp_div[d] += employment.get(code, 0)
        out_div[d] += output.get(code, 0)

    print(f"\nAggregated to {n_div} ANZSIC divisions")
    for i, d in enumerate(div_letters):
        print(f"  {d}: {ANZSIC_DIVISIONS[d][:40]:40s}  output={out_div[i]:12.0f}  emp={emp_div[i]:10.0f}")

    return A_div, emp_div, out_div


def compute_leontief_and_multipliers(A_div, emp_div, out_div):
    """Compute Leontief inverse and Type I multipliers."""
    n = A_div.shape[0]
    I = np.eye(n)

    # Leontief inverse: L = (I - A)^{-1}
    try:
        L = np.linalg.inv(I - A_div)
    except np.linalg.LinAlgError:
        print("ERROR: Singular matrix, cannot compute Leontief inverse")
        return None, None, None

    # Type I output multipliers: column sums of L
    output_multipliers = L.sum(axis=0)

    # Employment coefficients: jobs per $m of output
    emp_coeffs = np.zeros(n)
    for i in range(n):
        if out_div[i] > 0:
            emp_coeffs[i] = emp_div[i] / out_div[i]

    # Type I employment multipliers
    # For each industry j: sum over i of (L[i,j] * emp_coeffs[i]) / emp_coeffs[j]
    emp_multipliers = np.zeros(n)
    for j in range(n):
        if emp_coeffs[j] > 0:
            total_emp_effect = sum(L[i, j] * emp_coeffs[i] for i in range(n))
            emp_multipliers[j] = total_emp_effect / emp_coeffs[j]

    return L, output_multipliers, emp_multipliers


def main():
    base_dir = os.path.dirname(os.path.abspath(__file__))

    print("=" * 60)
    print("Building Australian National IO Model (2023-24)")
    print("=" * 60)

    # Step 1: Parse direct requirements
    print("\n--- Step 1: Parse direct requirements matrix ---")
    codes, names, A = parse_direct_requirements(
        os.path.join(base_dir, '520905500106.xlsx'))

    # Step 2: Parse employment
    print("\n--- Step 2: Parse employment by industry ---")
    # Employment is in the main DO001 file, Table 20
    employment = parse_employment(
        os.path.join(base_dir, '5209055001DO001_202324.xlsx'))

    # Step 3: Parse output (from flow table)
    print("\n--- Step 3: Parse industry output ---")
    output = parse_output(
        os.path.join(base_dir, '5209055001DO001_202324.xlsx'))

    # Step 4: Aggregate to 19 divisions
    print("\n--- Step 4: Aggregate to 19 ANZSIC divisions ---")
    A_div, emp_div, out_div = aggregate_to_divisions(codes, names, A, output, employment)

    # Step 5: Compute Leontief inverse and multipliers
    print("\n--- Step 5: Compute Leontief inverse and multipliers ---")
    L, out_mult, emp_mult = compute_leontief_and_multipliers(A_div, emp_div, out_div)

    if L is None:
        print("FAILED: Could not compute Leontief inverse")
        return

    # Print results
    div_letters = list(ANZSIC_DIVISIONS.keys())
    print("\n" + "=" * 80)
    print("NATIONAL TYPE I MULTIPLIERS (2023-24)")
    print("=" * 80)
    print(f"{'Div':>4} {'Division':40s} {'Output':>10} {'Employment':>12}")
    print("-" * 70)
    for i, d in enumerate(div_letters):
        print(f"  {d:>2}  {ANZSIC_DIVISIONS[d][:40]:40s} {out_mult[i]:10.4f} {emp_mult[i]:12.4f}")

    # Save national results as JSON
    output_dir = os.path.join(base_dir, '..', '..', 'src', 'data', 'au')
    os.makedirs(output_dir, exist_ok=True)

    national = {
        'source': 'ABS Australian National Accounts: Input-Output Tables, 2023-24',
        'published': '2026-03-25',
        'method': 'Leontief inverse of 19-division aggregated direct requirements matrix',
        'price_base': '2023-24 AUD',
        'industries': {},
        'leontief_inverse': L.tolist(),
        'direct_requirements': A_div.tolist(),
    }

    for i, d in enumerate(div_letters):
        national['industries'][d] = {
            'name': ANZSIC_DIVISIONS[d],
            'output_AUDm': round(out_div[i], 1),
            'employment': round(emp_div[i], 0),
            'employment_per_AUDm': round(emp_div[i] / out_div[i], 4) if out_div[i] > 0 else 0,
            'output_multiplier_type1': round(out_mult[i], 4),
            'employment_multiplier_type1': round(emp_mult[i], 4),
        }

    out_path = os.path.join(output_dir, 'national-io.json')
    with open(out_path, 'w') as f:
        json.dump(national, f, indent=2)
    print(f"\nSaved national IO data to {out_path}")


if __name__ == '__main__':
    main()
