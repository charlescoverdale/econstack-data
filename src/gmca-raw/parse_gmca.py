#!/usr/bin/env python3
"""Parse the GMCA Unit Cost Database Excel into structured JSON."""

import json
import openpyxl
import re

DATA_SHEETS = ['Crime', 'Education & Skills', 'Employment & Economy',
               'Fire', 'Health', 'Housing', 'Social Services']

def clean_text(val):
    """Clean cell value to a string."""
    if val is None:
        return ""
    s = str(val).strip()
    s = s.replace('\n', ' ').replace('\r', ' ')
    s = re.sub(r'\s+', ' ', s)
    return s

def parse_cost(val):
    """Parse a cost value to a number."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return round(float(val), 2)
    s = str(val).strip().replace(',', '').replace('£', '')
    s = re.sub(r'[^\d.\-]', '', s)
    if not s:
        return None
    try:
        return round(float(s), 2)
    except ValueError:
        return None

def main():
    wb = openpyxl.load_workbook('unit-cost-database.xlsx', read_only=True)

    database = {
        "source": "GMCA Unit Cost Database v1.4 (New Economy Manchester)",
        "publisher": "Greater Manchester Combined Authority Research Team",
        "url": "https://golab.bsg.ox.ac.uk/knowledge-bank/resources/unit-cost-database/",
        "license": "Creative Commons (via GO Lab, Oxford University)",
        "price_base": "Various (see individual entries). Most costs are 2013-14 or 2014-15 GBP.",
        "inflation_note": "Many costs are from 2013-15 publications. Apply GDP deflator to update to current prices. Approximate uplift to 2024-25: multiply by 1.30-1.35.",
        "total_entries": 0,
        "categories": {}
    }

    total = 0

    for sheet_name in DATA_SHEETS:
        if sheet_name not in wb.sheetnames:
            print(f"  SKIP: {sheet_name} not found")
            continue

        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        # Skip header rows (first 2 rows are headers)
        entries = []
        for row in rows[2:]:
            vals = list(row)
            if len(vals) < 8:
                continue

            category = clean_text(vals[0])
            detail = clean_text(vals[1])
            code = clean_text(vals[2])
            description = clean_text(vals[3])
            unit = clean_text(vals[4])
            agency_1 = clean_text(vals[5])
            agency_2 = clean_text(vals[6])
            cost = parse_cost(vals[7])

            if not code or not description:
                continue
            if cost is None:
                continue

            entry = {
                "code": code,
                "category": category if category else None,
                "detail": detail if detail else None,
                "description": description,
                "unit": unit if unit else None,
                "cost_gbp": cost,
                "agency": agency_1 if agency_1 else None,
                "agency_detail": agency_2 if agency_2 else None,
            }

            # Check for additional columns (source, notes)
            if len(vals) > 8:
                source = clean_text(vals[8]) if len(vals) > 8 else ""
                notes = clean_text(vals[9]) if len(vals) > 9 else ""
                if source:
                    entry["source_detail"] = source
                if notes:
                    entry["notes"] = notes

            entries.append(entry)

        # Fill forward category and detail from previous rows
        last_cat = ""
        last_detail = ""
        for e in entries:
            if e["category"]:
                last_cat = e["category"]
            else:
                e["category"] = last_cat
            if e["detail"]:
                last_detail = e["detail"]
            else:
                e["detail"] = last_detail

        cat_key = sheet_name.lower().replace(' & ', '_').replace(' ', '_')
        database["categories"][cat_key] = {
            "sheet": sheet_name,
            "count": len(entries),
            "entries": entries
        }
        total += len(entries)
        print(f"  {sheet_name}: {len(entries)} entries")

    database["total_entries"] = total
    print(f"\nTotal: {total} entries")

    wb.close()

    # Save
    out_path = '../../parameters/uk/gmca-unit-costs-full.json'
    with open(out_path, 'w') as f:
        json.dump(database, f, indent=2, ensure_ascii=False)
    print(f"Saved to {out_path}")

    # Also print some stats
    for cat_key, cat_data in database["categories"].items():
        subcats = set()
        for e in cat_data["entries"]:
            if e.get("detail"):
                subcats.add(e["detail"])
        print(f"  {cat_key}: {cat_data['count']} entries, {len(subcats)} subcategories")

if __name__ == '__main__':
    main()
